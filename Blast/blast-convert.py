#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Name: blast-convert.pdf
# Description: Reads a blast table and re-writes it as a GTF file.

##
# MODULES
##

import openpyxl
import sys, re, argparse
from typing import List, Dict


def main(args):

    if args.type == 'xlsx':
        xlsx = True
        wb = openpyxl.load_workbook(filename=args.input, data_only=True)
        first_sheet = wb[wb.sheetnames[0]]
        columns = first_sheet.max_column
        rows = first_sheet.max_row
        blastFields: List[List[any]] = []
        for i in range(rows)[1:]:
            entry: List[any] = []
            for j in range(columns):
                entry.append(first_sheet.cell(row=i+1, column=j+1).value)
            blastFields.append(entry)  
        for i in blastFields:
            print(i)
    elif args.type == 'tsv':
        TSV = True
        blastFields: List[List[str]] = []
        for line in open(args.input, 'r', encoding="utf8"):
            columns = line.strip().split("\t")
            blastFields.append(columns) 
        for i in blastFields:
            print(i)
    else:
        if args.type == 'csv':
            CSV = True
            blastFields : List[List[str]] = []
            for line in open(args.input, 'r', encoding="utf8"):
                columns = line.strip().split(",")
                blastFields.append(columns) 
            for i in blastFields:
                print(i)

    output = open(f'converted_{args.input}.gtf', 'w')
    for column in blastFields:
        if TSV or CSV:

            attr: str = 'cluster="{}"; pident="{}"; length="{}"; mismatch="{}"; gapopen="{}"; bitscore="{}", qstart="{}": qend="{}";'.format(column[1], column[2], column[3], column[4], column[5], column[6], column[7], column[11])
            seqid, source, feature, start, end, score, strand, frame = column[0], 'BLAST', 'CDS', column[8], column[9] ,column[10], '.', '.'
        else:
            attr: str = 'gene="{}"; cluster="{}"; pident="{}"; length="{}"; mismatch="{}"; gapopen="{}"; bitscore="{}", qstart="{}": qend="{}";'.format(column[0], column[2], column[3], column[4], column[5], column[6], column[7], column[8], column[12])
            seqid, source, feature, start, end, score, strand, frame = column[1], 'BLAST', 'CDS', column[9], column[10] ,column[11], '.', '.'
        output.write('{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n'.format(seqid, source, feature, start, end, score, strand, frame, attr))
    output.close()
if __name__ == '__main__':
    	# The arguments are defined in the next few lines

	parser=argparse.ArgumentParser(description='Converts blast format 6 to GTF format.')
	parser.add_argument("-i", "--input", help = "Blast File Name", type = str)
	parser.add_argument("-t", "--type", help = "File type. tsv, csv or xlsx. Targets first excel spreadsheet.", type = str, required = False, choices=['tsv', 'csv', 'xlsx'], default='TSV')
	args = parser.parse_args()
	main(args) # This will call the main fuction

exit() # Quit this script